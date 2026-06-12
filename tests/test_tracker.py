"""Tests for `utils/tracker`."""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook

from utils.tracker import (
    add_application,
    application_exists,
    delete_application,
    edit_application,
    get_followups_due,
    show_tracker,
    update_status,
)


# ---------------------------------------------------------------------------
# add_application
# ---------------------------------------------------------------------------


def test_add_application_date_logic(tmp_path: Path) -> None:
    """Appends a row with today's date and follow-up date +14 days."""
    path = tmp_path / "tracker.xlsx"
    add_application(path, "Acme", "Engineer", "LinkedIn", 80, "note")

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 2  # header + 1 data row
    _, _, date_applied, _, _, status, follow_up, notes = rows[1]
    assert date_applied == date.today().isoformat()
    assert follow_up == (date.today() + timedelta(days=14)).isoformat()
    assert status == "Applied"


@pytest.mark.parametrize("company,role", [("", "Eng"), ("Acme", ""), ("  ", "Dev")])
def test_add_application_empty_validation(tmp_path: Path, company: str, role: str) -> None:
    """Raise ValueError when company or role is empty or whitespace-only."""
    path = tmp_path / "tracker.xlsx"
    with pytest.raises(ValueError):
        add_application(path, company, role, "LinkedIn", 80, "")


# ---------------------------------------------------------------------------
# application_exists
# ---------------------------------------------------------------------------


def test_application_exists_case_insensitive(tmp_tracker: Path) -> None:
    """Return True for exact match ignoring case."""
    assert application_exists(tmp_tracker, "acme", "engineer") is True
    assert application_exists(tmp_tracker, "ACME", "ENGINEER") is True


def test_application_exists_missing_file(tmp_path: Path) -> None:
    """Return False when the tracker file does not exist."""
    path = tmp_path / "nonexistent.xlsx"
    assert application_exists(path, "Acme", "Engineer") is False


# ---------------------------------------------------------------------------
# delete_application
# ---------------------------------------------------------------------------


def test_delete_application_removes_row(tmp_tracker: Path) -> None:
    """Remove the correct row and return True."""
    assert delete_application(tmp_tracker, "Acme", "Engineer") is True
    assert application_exists(tmp_tracker, "Acme", "Engineer") is False


def test_delete_application_not_found(tmp_tracker: Path) -> None:
    """Return False when the row is not found."""
    assert delete_application(tmp_tracker, "Unknown", "Role") is False


# ---------------------------------------------------------------------------
# edit_application
# ---------------------------------------------------------------------------


def test_edit_application_updates_field(tmp_tracker: Path) -> None:
    """Update the correct field in the correct row."""
    assert edit_application(tmp_tracker, "Acme", "Engineer", "Notes", "Phone screen done") is True

    from openpyxl import load_workbook

    wb = load_workbook(tmp_tracker, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # Notes is column index 8 (0-based)
    assert rows[1][7] == "Phone screen done"


def test_edit_application_non_editable_field(tmp_tracker: Path) -> None:
    """Return False when the field is not in the editable set."""
    assert edit_application(tmp_tracker, "Acme", "Engineer", "Company", "NewCo") is False


# ---------------------------------------------------------------------------
# update_status
# ---------------------------------------------------------------------------


def test_update_status_valid(tmp_tracker: Path) -> None:
    """Change the status for a matching row."""
    update_status(tmp_tracker, "Acme", "Engineer", "Interviewing")

    from openpyxl import load_workbook

    wb = load_workbook(tmp_tracker, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[1][5] == "Interviewing"


def test_update_status_invalid_status(tmp_path: Path) -> None:
    """Raise ValueError for an invalid status string."""
    with pytest.raises(ValueError, match="Invalid status"):
        update_status(tmp_path / "tracker.xlsx", "Acme", "Engineer", "Ghosted")


# ---------------------------------------------------------------------------
# get_followups_due
# ---------------------------------------------------------------------------


def test_get_followups_due_filters_and_dates(tmp_tracker: Path) -> None:
    """Return only Applied rows whose follow-up date is today or earlier."""
    # Create an additional row with a past follow-up date
    add_application(tmp_tracker, "OldCo", "Dev", "Indeed", 70, "")
    edit_application(tmp_tracker, "OldCo", "Dev", "Follow-up Date", "2000-01-01")

    due = get_followups_due(tmp_tracker)
    assert len(due) == 1
    assert due[0]["Company"] == "OldCo"
    assert due[0]["Role"] == "Dev"


# ---------------------------------------------------------------------------
# show_tracker
# ---------------------------------------------------------------------------


def test_show_tracker_empty(tmp_path: Path) -> None:
    """Print the empty-tracker message without crashing."""
    path = tmp_path / "tracker.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    headers = [
        "Company",
        "Role",
        "Date Applied",
        "Source",
        "Match Score",
        "Status",
        "Follow-up Date",
        "Notes",
        "Cover Letter Path",
    ]
    ws.append(headers)
    wb.save(path)

    with patch("utils.tracker.console.print") as mock_print:
        show_tracker(path)
    mock_print.assert_called_once()
    assert "No applications found" in mock_print.call_args[0][0]


def test_show_tracker_with_orphaned_column(tmp_path: Path) -> None:
    """Render an old tracker that still has the 'Cover Letter Path' column."""
    path = tmp_path / "tracker.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    headers = [
        "Company",
        "Role",
        "Date Applied",
        "Source",
        "Match Score",
        "Status",
        "Follow-up Date",
        "Notes",
        "Cover Letter Path",
    ]
    ws.append(headers)
    ws.append([
        "OldCo",
        "Engineer",
        "2024-01-01",
        "Indeed",
        "90",
        "Applied",
        "2024-01-15",
        "Note here",
        "/tmp/cover_letter.txt",
    ])
    wb.save(path)

    with patch("utils.tracker.console.print") as mock_print:
        show_tracker(path)
    mock_print.assert_called_once()
    output = mock_print.call_args[0][0]
    # output is a Rich Table object; verify it has the orphaned column
    from rich.table import Table
    assert isinstance(output, Table)
    assert len(output.columns) == 9  # includes orphaned "Cover Letter Path"
