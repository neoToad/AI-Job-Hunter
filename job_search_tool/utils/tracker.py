"""Application tracker backed by an Excel workbook (openpyxl).

Provides CRUD-style helpers for logging job applications, detecting duplicates,
and surfacing follow-ups that are due.
"""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from rich.table import Table

from utils.console import console

_HEADERS = [
    "Company",
    "Role",
    "Date Applied",
    "Source",
    "Match Score",
    "Status",
    "Follow-up Date",
    "Notes",
    "Cover Letter Path",
]

# openpyxl pattern: use read_only=True for pure reads (faster, safer) and
# read_only=False (the default) for any mutation so the workbook can be saved.


def _safe_save(wb: Workbook, path: Path) -> None:
    """Save *wb* to *path*, translating `PermissionError` into a friendly message.

    Args:
        wb: The workbook to save.
        path: Destination file path.

    Raises:
        PermissionError: With a user-friendly message if the file is locked.
    """
    try:
        wb.save(path)
    except PermissionError as exc:
        raise PermissionError(
            "Tracker file may be open in another program. Close it and try again."
        ) from exc


def _get_or_create_workbook(path: Path) -> Workbook:
    """Load an existing workbook or create a new one with bold headers.

    Args:
        path: Path to the .xlsx tracker file.

    Returns:
        An openpyxl Workbook instance (loaded or freshly created).
    """
    if path.exists():
        try:
            wb = load_workbook(path, read_only=False)
            ws = wb.active
            # Ensure headers are present; if sheet is empty or malformed, recreate.
            if ws is None or ws.max_row == 0:
                raise ValueError("Empty worksheet")
            return wb
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.title = "Applications"
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"

    ws.append(_HEADERS)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    return wb


def application_exists(path: Path, company: str, role: str) -> bool:
    """Return True if an application for the same *company* + *role*
    (case-insensitive, stripped) already exists in the tracker.

    Args:
        path: Path to the .xlsx tracker file.
        company: Company name.
        role: Job role/title.

    Returns:
        Boolean indicating whether a duplicate entry was found.
    """
    if not path.exists():
        return False

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return False

    target_company = company.strip().lower()
    target_role = role.strip().lower()

    # Skip header row
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 2:
            continue
        if str(row[0] or "").strip().lower() == target_company and str(
            row[1] or ""
        ).strip().lower() == target_role:
            return True
    return False


def add_application(
    path: Path,
    company: str,
    role: str,
    source: str,
    match_score: float | int | str,
    notes: str,
    cover_letter_path: str | Path,
) -> None:
    """Append a new application row to the tracker.

    Auto-fills:
    - *Date Applied* → today's date (ISO format).
    - *Status* → ``"Applied"``.
    - *Follow-up Date* → 14 days from today.

    Args:
        path: Path to the .xlsx tracker file.
        company: Company name.
        role: Job role/title.
        source: Where the posting was found (e.g. LinkedIn, Indeed).
        match_score: Numerical or string match score.
        notes: Free-text notes.
        cover_letter_path: Path to the generated cover letter.

    Raises:
        ValueError: If *company* or *role* is empty or whitespace-only.
    """
    if not company or not company.strip():
        raise ValueError("Company name must be a non-empty string.")
    if not role or not role.strip():
        raise ValueError("Role must be a non-empty string.")

    wb = _get_or_create_workbook(path)
    ws = wb.active
    today = date.today()
    follow_up = today + timedelta(days=14)

    ws.append(
        [
            company,
            role,
            today.isoformat(),
            source,
            str(match_score),
            "Applied",
            follow_up.isoformat(),
            notes,
            str(cover_letter_path),
        ]
    )
    _safe_save(wb, path)


def show_tracker(path: Path) -> None:
    """Print every row in the tracker as a formatted Rich table.

    Args:
        path: Path to the .xlsx tracker file.
    """
    if not path.exists():
        print(f"Tracker not found: {path}")
        return

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        print("Tracker workbook has no active sheet.")
        return

    table = Table(title="Job Application Tracker", show_header=True, header_style="bold magenta")
    for header in _HEADERS:
        table.add_column(header)

    row_count = 0
    for row in ws.iter_rows(values_only=True):
        # Render every cell as a string; empty cells become ""
        rendered = [str(cell or "") for cell in row]
        if any(rendered):
            table.add_row(*rendered)
            row_count += 1

    if row_count == 0:
        console.print("[yellow]No applications found.[/]")
    else:
        console.print(table)


def get_followups_due(path: Path) -> list[dict[str, Any]]:
    """Return applications whose Status is ``"Applied"`` and whose
    Follow-up Date is today or earlier.

    Args:
        path: Path to the .xlsx tracker file.

    Returns:
        List of application rows as dictionaries keyed by column header.
    """
    if not path.exists():
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []

    today = date.today()
    due: list[dict[str, Any]] = []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h or f"Col{i}") for i, h in enumerate(rows[0])]
    for row in rows[1:]:
        row_dict = dict(zip(headers, row))
        status = str(row_dict.get("Status", "")).strip()
        follow_up_raw = row_dict.get("Follow-up Date", "")
        try:
            follow_up = date.fromisoformat(str(follow_up_raw).strip())
        except (ValueError, TypeError):
            continue

        if status == "Applied" and follow_up <= today:
            due.append(row_dict)

    return due


_VALID_STATUSES = {"Applied", "Interviewing", "Offer", "Rejected", "Withdrawn"}


def update_status(path: Path, company: str, role: str, new_status: str) -> None:
    """Update the Status column for the matching company + role row.

    Args:
        path: Path to the .xlsx tracker file.
        company: Company name to match.
        role: Role title to match.
        new_status: One of ``"Applied"``, ``"Interviewing"``, ``"Offer"``,
            ``"Rejected"``, ``"Withdrawn"``.

    Raises:
        ValueError: If *new_status* is not a valid status.
        FileNotFoundError: If the tracker file does not exist.
    """
    if new_status not in _VALID_STATUSES:
        raise ValueError(
            f"Invalid status '{new_status}'. Must be one of: {', '.join(sorted(_VALID_STATUSES))}"
        )

    if not path.exists():
        raise FileNotFoundError(f"Tracker not found at {path}")

    # Open read-write so we can save changes back to disk.
    wb = load_workbook(path, read_only=False)
    ws = wb.active
    if ws is None:
        raise ValueError("Tracker workbook has no active sheet.")

    target_company = company.strip().lower()
    target_role = role.strip().lower()
    status_col: int | None = None

    # Find the Status column index from the header row
    for idx, cell in enumerate(ws[1], start=1):
        if str(cell.value or "").strip().lower() == "status":
            status_col = idx
            break

    if status_col is None:
        raise ValueError("Tracker is missing a 'Status' column.")

    updated = False
    for row in ws.iter_rows(min_row=2):
        if len(row) < 2:
            continue
        row_company = str(row[0].value or "").strip().lower()
        row_role = str(row[1].value or "").strip().lower()
        if row_company == target_company and row_role == target_role:
            row[status_col - 1].value = new_status
            updated = True
            break

    if not updated:
        raise ValueError(
            f"No application found for {company} — {role}."
        )

    _safe_save(wb, path)


_EDITABLE_FIELDS = {
    "Source",
    "Match Score",
    "Status",
    "Follow-up Date",
    "Notes",
    "Cover Letter Path",
}


def delete_application(path: Path, company: str, role: str) -> bool:
    """Delete the row matching *company* + *role* (case-insensitive).

    Args:
        path: Path to the .xlsx tracker file.
        company: Company name to match.
        role: Role title to match.

    Returns:
        ``True`` if a row was found and deleted, ``False`` otherwise.
    """
    if not path.exists():
        return False

    # Open read-write so we can save changes back to disk.
    wb = load_workbook(path, read_only=False)
    ws = wb.active
    if ws is None:
        return False

    target_company = company.strip().lower()
    target_role = role.strip().lower()

    deleted = False
    for row in ws.iter_rows(min_row=2):
        if len(row) < 2:
            continue
        row_company = str(row[0].value or "").strip().lower()
        row_role = str(row[1].value or "").strip().lower()
        if row_company == target_company and row_role == target_role:
            ws.delete_rows(row[0].row)
            deleted = True
            break

    if deleted:
        _safe_save(wb, path)
    return deleted


def edit_application(path: Path, company: str, role: str, field: str, value: str) -> bool:
    """Update a single column for the row matching *company* + *role*.

    Args:
        path: Path to the .xlsx tracker file.
        company: Company name to match.
        role: Role title to match.
        field: Column header to update. Must be one of the editable fields.
        value: New value for the cell.

    Returns:
        ``True`` if the row and field were found and updated, ``False`` otherwise.
    """
    if not path.exists():
        return False

    if field not in _EDITABLE_FIELDS:
        return False

    # Open read-write so we can save changes back to disk.
    wb = load_workbook(path, read_only=False)
    ws = wb.active
    if ws is None:
        return False

    target_company = company.strip().lower()
    target_role = role.strip().lower()

    # Find the column index for the requested field
    field_col: int | None = None
    for idx, cell in enumerate(ws[1], start=1):
        if str(cell.value or "").strip() == field:
            field_col = idx
            break

    if field_col is None:
        return False

    updated = False
    for row in ws.iter_rows(min_row=2):
        if len(row) < 2:
            continue
        row_company = str(row[0].value or "").strip().lower()
        row_role = str(row[1].value or "").strip().lower()
        if row_company == target_company and row_role == target_role:
            row[field_col - 1].value = value
            updated = True
            break

    if updated:
        _safe_save(wb, path)
    return updated